"""Parse an IELTS test Excel workbook and build preview/confirm structures."""

from __future__ import annotations

import io
from dataclasses import dataclass, field

import openpyxl

# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------


@dataclass
class ParsedQuestion:
    order: int
    question_type: str
    question: str
    options: list[str]       # already normalised to "A. text" form
    answer: str              # raw cell value
    instruction: str | None
    # writing-only fields
    task_number: int | None = None
    min_words: int | None = None
    # listening-only
    part: int | None = None
    # grouping
    group: int | None = None


@dataclass
class ParsedSection:
    sheet_name: str
    section_kind: str        # "reading", "writing", "listening"
    passage: str | None      # reading passage or listening audioscript
    audio_filename: str | None
    questions: list[ParsedQuestion] = field(default_factory=list)
    # listening part number (1-4)
    part_number: int | None = None


@dataclass
class ParsedTest:
    title: str
    description: str | None
    type: str                # "academic" | "general"
    sections: list[ParsedSection] = field(default_factory=list)


@dataclass
class SectionSummary:
    sheet_name: str
    kind: str
    passage_word_count: int | None
    questions_count: int
    tasks_count: int | None
    audio_filename: str | None


@dataclass
class PreviewResult:
    title: str
    description: str | None
    type: str
    sections: list[SectionSummary]
    total_questions: int
    warnings: list[str]
    errors: list[str]


# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------

_TYPE_MAP: dict[str, str] = {
    "tf": "true_false_ng",
    "tfng": "true_false_ng",
    "true_false": "true_false_ng",
    "true/false/ng": "true_false_ng",
    "true_false_not_given": "true_false_ng",
    "gap": "gap_fill",
    "fill": "gap_fill",
    "completion": "gap_fill",
    "gap_fill": "gap_fill",
    "gap fill": "gap_fill",
    "mc": "mcq",
    "multiple_choice": "mcq",
    "multiple choice": "mcq",
    "mcq": "mcq",
    "match": "matching",
    "matching": "matching",
    "headings": "matching_headings",
    "matching_headings": "matching_headings",
    "information": "matching_information",
    "matching_information": "matching_information",
    "features": "matching_features",
    "matching_features": "matching_features",
    "task1": "task1",
    "essay": "essay",
    # yes/no/not given
    "yn": "yes_no_ng",
    "ynng": "yes_no_ng",
    "yes_no": "yes_no_ng",
    "yes_no_ng": "yes_no_ng",
    "yes/no/ng": "yes_no_ng",
    "yes/no/not given": "yes_no_ng",
    # sentence completion
    "sc": "sentence_completion",
    "sentence": "sentence_completion",
    "sent_completion": "sentence_completion",
    "sentence_completion": "sentence_completion",
    # short answer
    "sa": "short_answer",
    "short": "short_answer",
    "short_ans": "short_answer",
    "short_answer": "short_answer",
}

_ALLOWED_QUESTION_TYPES = {
    "mcq", "gap_fill", "matching", "true_false_ng",
    "matching_headings", "matching_information", "matching_features",
    "yes_no_ng", "sentence_completion", "short_answer",
}
_ALLOWED_WRITING_TYPES = {"task1", "essay"}


def _normalize_type(raw: str) -> str:
    return _TYPE_MAP.get(raw.strip().lower(), raw.strip().lower())


def _parse_options(raw: str | None) -> list[str]:
    """Split 'A.xxx;B.yyy' into ['A. xxx', 'B. yyy']."""
    if not raw:
        return []
    parts: list[str] = []
    for seg in raw.split(";"):
        seg = seg.strip()
        if not seg:
            continue
        # normalise "A.xxx" -> "A. xxx" or "A. xxx" -> "A. xxx"
        if len(seg) >= 2 and seg[1] == "." and not seg.startswith("i"):
            letter = seg[0].upper()
            rest = seg[2:].strip()
            parts.append(f"{letter}. {rest}")
        else:
            # roman numeral or bare text — keep as-is
            parts.append(seg)
    return parts


_LETTER_TO_INDEX = {
    "A": 0, "B": 1, "C": 2, "D": 3, "E": 4,
    "F": 5, "G": 6, "H": 7, "I": 8, "J": 9,
}

_ROMAN_TO_INDEX = {
    "i": 0, "ii": 1, "iii": 2, "iv": 3, "v": 4,
    "vi": 5, "vii": 6, "viii": 7, "ix": 8, "x": 9,
}


def _resolve_letter(answer: str, options: list[str]) -> str | None:
    """Given answer letter (e.g. 'B' or 'ii') return the matching option string."""
    a = answer.strip()
    idx = _LETTER_TO_INDEX.get(a.upper())
    if idx is None:
        idx = _ROMAN_TO_INDEX.get(a.lower())
    if idx is not None and idx < len(options):
        return options[idx]
    return None


def _build_content(qtype: str, question: str, options: list[str], instruction: str | None) -> dict:
    if qtype in ("true_false_ng", "yes_no_ng"):
        d: dict = {"statement": question}
        if instruction:
            d["instruction"] = instruction
        return d
    if qtype == "gap_fill":
        d = {"text": question}
        if instruction:
            d["instruction"] = instruction
        return d
    if qtype in ("sentence_completion", "short_answer"):
        # options column may carry max_words as a number
        max_words = 3
        if options:
            try:
                max_words = int(options[0])
            except (ValueError, TypeError):
                max_words = 3
        d = {"prompt": question, "max_words": max_words}
        if instruction:
            d["instruction"] = instruction
        return d
    if qtype == "mcq":
        d = {"question": question, "options": options}
        if instruction:
            d["instruction"] = instruction
        return d
    if qtype == "matching":
        d = {"items": [question], "options": options}
        if instruction:
            d["instruction"] = instruction
        return d
    if qtype in {"matching_headings", "matching_information", "matching_features"}:
        # Options are hoisted to QuestionGroup.options_shared; only the statement/label goes in content
        d = {"question": question}
        if instruction:
            d["instruction"] = instruction
        return d
    if qtype in {"task1", "essay"}:
        d = {"prompt": question}
        if instruction:
            d["instruction"] = instruction
        return d
    return {"question": question}


def _build_answer_key(qtype: str, answer: str, options: list[str]) -> dict | None:
    if not answer:
        return None
    if qtype == "true_false_ng":
        mapping = {
            "true": "True",
            "false": "False",
            "not given": "Not Given",
            "ng": "Not Given",
        }
        return {"correct": mapping.get(answer.strip().lower(), answer.strip())}
    if qtype == "yes_no_ng":
        mapping = {
            "yes": "Yes",
            "no": "No",
            "not given": "Not Given",
            "ng": "Not Given",
            "not_given": "Not Given",
        }
        return {"correct": mapping.get(answer.strip().lower(), answer.strip())}
    if qtype in ("sentence_completion", "short_answer"):
        ans = answer.strip()
        if ";" in ans:
            variants = [v.strip() for v in ans.split(";") if v.strip()]
            return {"correct": variants}
        return {"correct": ans}
    if qtype == "gap_fill":
        ans = answer.strip()
        if ";" in ans:
            variants = [v.strip() for v in ans.split(";") if v.strip()]
            return {"correct": variants}
        return {"correct": ans}
    if qtype == "mcq":
        resolved = _resolve_letter(answer, options)
        if resolved:
            return {"correct": resolved}
        return {"correct": answer.strip()}
    if qtype == "matching":
        resolved = _resolve_letter(answer, options)
        if resolved:
            return {"correct": resolved}
        return {"correct": answer.strip()}
    if qtype in {"matching_headings", "matching_information", "matching_features"}:
        # Store bare prefix as-is (e.g. "iii", "A") — do NOT resolve to full option string
        return {"correct": answer.strip()}
    if qtype in {"task1", "essay"}:
        return None
    return None


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def _str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _int_or_none(val) -> int | None:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _parse_info_sheet(ws) -> tuple[str, str | None, str]:
    title = ""
    description: str | None = None
    test_type = "academic"
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        key = _str(row[0]).lower()
        val = _str(row[1]) if len(row) > 1 else ""
        if key == "title":
            title = val
        elif key == "description":
            description = val or None
        elif key == "type":
            test_type = val.lower() if val.lower() in {"academic", "general"} else "academic"
    return title, description, test_type


def _has_group_column(ws, header_row: int) -> bool:
    """Check if the sheet has the new 'group' column at position 2 (index 1)."""
    headers = [_str(ws.cell(row=header_row, column=c).value).lower() for c in range(1, 4)]
    return "group" in headers


def _parse_reading_sheet(ws, sheet_name: str) -> ParsedSection:
    passage = _str(ws["A1"].value)
    questions: list[ParsedQuestion] = []
    orders_seen: set[int] = set()
    warnings: list[str] = []

    has_group = _has_group_column(ws, header_row=3)

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or all(v is None for v in row):
            continue

        if has_group:
            order_raw, group_raw, type_raw, question_raw, opts_raw, ans_raw, inst_raw = (
                (row[i] if i < len(row) else None) for i in range(7)
            )
            group_num = _int_or_none(group_raw)
        else:
            order_raw, type_raw, question_raw, opts_raw, ans_raw, inst_raw = (
                (row[i] if i < len(row) else None) for i in range(6)
            )
            group_num = None

        if order_raw is None and question_raw is None:
            continue
        order = _int_or_none(order_raw) or 0
        qtype = _normalize_type(_str(type_raw))
        question = _str(question_raw)
        options = _parse_options(_str(opts_raw))
        answer = _str(ans_raw)
        instruction = _str(inst_raw) or None

        if order in orders_seen:
            warnings.append(f"{sheet_name}: duplicate order {order}")
        orders_seen.add(order)

        questions.append(
            ParsedQuestion(
                order=order,
                question_type=qtype,
                question=question,
                options=options,
                answer=answer,
                instruction=instruction,
                group=group_num,
            )
        )

    return ParsedSection(
        sheet_name=sheet_name,
        section_kind="reading",
        passage=passage or None,
        audio_filename=None,
        questions=questions,
    )


def _parse_writing_sheet(ws, sheet_name: str) -> ParsedSection:
    questions: list[ParsedQuestion] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        order_raw, task_num_raw, type_raw, prompt_raw, inst_raw, min_words_raw = (
            (row[i] if i < len(row) else None) for i in range(6)
        )
        if prompt_raw is None:
            continue
        order = _int_or_none(order_raw) or 0
        qtype = _normalize_type(_str(type_raw))
        if qtype not in _ALLOWED_WRITING_TYPES:
            qtype = "essay"
        questions.append(
            ParsedQuestion(
                order=order,
                question_type=qtype,
                question=_str(prompt_raw),
                options=[],
                answer="",
                instruction=_str(inst_raw) or None,
                task_number=_int_or_none(task_num_raw),
                min_words=_int_or_none(min_words_raw),
            )
        )
    return ParsedSection(
        sheet_name=sheet_name,
        section_kind="writing",
        passage=None,
        audio_filename=None,
        questions=questions,
    )


def _parse_listening_sheet(ws, sheet_name: str) -> ParsedSection:
    audio_filename = _str(ws["A1"].value) or None
    audioscript = _str(ws["A2"].value) or None
    questions: list[ParsedQuestion] = []
    orders_seen: set[int] = set()

    has_group = _has_group_column(ws, header_row=4)

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or all(v is None for v in row):
            continue

        if has_group:
            order_raw, group_raw, type_raw, question_raw, opts_raw, ans_raw, part_raw = (
                (row[i] if i < len(row) else None) for i in range(7)
            )
            group_num = _int_or_none(group_raw)
        else:
            order_raw, type_raw, question_raw, opts_raw, ans_raw, part_raw = (
                (row[i] if i < len(row) else None) for i in range(6)
            )
            group_num = None

        if order_raw is None and question_raw is None:
            continue
        order = _int_or_none(order_raw) or 0
        qtype = _normalize_type(_str(type_raw))
        options = _parse_options(_str(opts_raw))
        answer = _str(ans_raw)

        if order in orders_seen:
            pass  # just continue; preview validation handles this
        orders_seen.add(order)

        questions.append(
            ParsedQuestion(
                order=order,
                question_type=qtype,
                question=_str(question_raw),
                options=options,
                answer=answer,
                instruction=None,
                part=_int_or_none(part_raw),
                group=group_num,
            )
        )

    # Infer part number from sheet name "Listening 1" -> 1
    part_num: int | None = None
    tail = sheet_name.replace("Listening", "").strip()
    try:
        part_num = int(tail)
    except ValueError:
        pass

    return ParsedSection(
        sheet_name=sheet_name,
        section_kind="listening",
        passage=audioscript,
        audio_filename=audio_filename,
        questions=questions,
        part_number=part_num,
    )


# ---------------------------------------------------------------------------
# Top-level parse
# ---------------------------------------------------------------------------

def parse_xlsx(file_bytes: bytes) -> ParsedTest:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    title = ""
    description: str | None = None
    test_type = "academic"
    sections: list[ParsedSection] = []

    for ws in wb.worksheets:
        name = ws.title
        if name == "Test Info":
            title, description, test_type = _parse_info_sheet(ws)
        elif name.startswith("Reading"):
            sections.append(_parse_reading_sheet(ws, name))
        elif name == "Writing":
            sections.append(_parse_writing_sheet(ws, name))
        elif name.startswith("Listening"):
            sections.append(_parse_listening_sheet(ws, name))

    return ParsedTest(
        title=title,
        description=description,
        type=test_type,
        sections=sections,
    )


# ---------------------------------------------------------------------------
# Preview builder
# ---------------------------------------------------------------------------

_IELTS_MAX_LISTENING = 4
_IELTS_MAX_READING = 3


def build_preview(parsed: ParsedTest) -> PreviewResult:
    warnings: list[str] = []
    errors: list[str] = []
    summaries: list[SectionSummary] = []
    total_questions = 0

    if not parsed.title:
        errors.append("Test Info: 'title' is required.")

    if parsed.type == "general":
        errors.append(
            "General Training is not supported yet. Change the 'type' cell to 'academic' and re-upload."
        )

    # IELTS count warnings
    listening_count = sum(1 for s in parsed.sections if s.section_kind == "listening")
    reading_count = sum(1 for s in parsed.sections if s.section_kind == "reading")
    if listening_count > _IELTS_MAX_LISTENING:
        warnings.append(
            f"Workbook has {listening_count} Listening sheets; IELTS standard is {_IELTS_MAX_LISTENING}. "
            f"Only the first {_IELTS_MAX_LISTENING} will be imported."
        )
    if reading_count > _IELTS_MAX_READING:
        warnings.append(
            f"Workbook has {reading_count} Reading sheets; IELTS standard is {_IELTS_MAX_READING}. "
            f"Only the first {_IELTS_MAX_READING} will be imported."
        )

    for sec in parsed.sections:
        q_count = len(sec.questions)
        total_questions += q_count

        # --- per-question validation ---
        orders: list[int] = []
        for q in sec.questions:
            orders.append(q.order)

            if q.question_type not in _ALLOWED_QUESTION_TYPES | _ALLOWED_WRITING_TYPES:
                warnings.append(
                    f"{sec.sheet_name} q{q.order}: unknown type '{q.question_type}'"
                )

            if sec.section_kind != "writing" and not q.answer:
                errors.append(f"{sec.sheet_name} q{q.order}: missing answer.")

            if q.question_type == "mcq" and not q.options:
                errors.append(f"{sec.sheet_name} q{q.order}: MCQ missing options.")

            if q.question_type == "matching_headings" and q.options and len(q.options) < 3:
                warnings.append(
                    f"{sec.sheet_name} q{q.order}: matching_headings should have at least 3 options."
                )
            if q.question_type == "matching_features" and q.options and len(q.options) < 2:
                warnings.append(
                    f"{sec.sheet_name} q{q.order}: matching_features should have at least 2 options."
                )

        dup_orders = {o for o in orders if orders.count(o) > 1}
        for d in sorted(dup_orders):
            warnings.append(f"{sec.sheet_name}: duplicate order {d}.")

        # --- section-level validation ---
        if sec.section_kind == "reading":
            if not sec.passage:
                errors.append(f"{sec.sheet_name}: passage text is missing.")
            else:
                wc = len(sec.passage.split())
                if wc < 800:
                    warnings.append(
                        f"{sec.sheet_name}: passage is {wc} words (recommend ≥ 800)."
                    )
            summaries.append(
                SectionSummary(
                    sheet_name=sec.sheet_name,
                    kind="reading",
                    passage_word_count=len(sec.passage.split()) if sec.passage else 0,
                    questions_count=q_count,
                    tasks_count=None,
                    audio_filename=None,
                )
            )

        elif sec.section_kind == "writing":
            # Validate writing task numbers and types
            for q in sec.questions:
                expected_task = q.task_number
                if expected_task is not None and expected_task not in (1, 2):
                    errors.append(
                        f"{sec.sheet_name} q{q.order}: task_number must be 1 or 2, got {expected_task}."
                    )
                # Normalise task number from order if missing
                if expected_task is None and q.order in (1, 2):
                    expected_task = q.order
                # Check type consistency
                if expected_task == 1 and q.question_type not in ("task1", "essay"):
                    warnings.append(
                        f"{sec.sheet_name} q{q.order}: Task 1 type should be 'task1'; got '{q.question_type}'. Will save as 'essay'."
                    )
                if expected_task == 2 and q.question_type not in ("essay",):
                    warnings.append(
                        f"{sec.sheet_name} q{q.order}: Task 2 type should be 'essay'; got '{q.question_type}'. Will save as 'essay'."
                    )

            summaries.append(
                SectionSummary(
                    sheet_name=sec.sheet_name,
                    kind="writing",
                    passage_word_count=None,
                    questions_count=q_count,
                    tasks_count=q_count,
                    audio_filename=None,
                )
            )

        elif sec.section_kind == "listening":
            if not sec.audio_filename:
                warnings.append(f"{sec.sheet_name}: no audio_filename in A1.")
            summaries.append(
                SectionSummary(
                    sheet_name=sec.sheet_name,
                    kind="listening",
                    passage_word_count=None,
                    questions_count=q_count,
                    tasks_count=None,
                    audio_filename=sec.audio_filename,
                )
            )

    return PreviewResult(
        title=parsed.title,
        description=parsed.description,
        type=parsed.type,
        sections=summaries,
        total_questions=total_questions,
        warnings=warnings,
        errors=errors,
    )
