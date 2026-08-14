"""Build the downloadable IELTS test template workbook."""

from __future__ import annotations

import io

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

_READING_TAB  = "2E7D32"
_WRITING_TAB  = "1565C0"
_LISTENING_TAB = "EF6C00"
_INFO_TAB     = "607D8B"

_AUTHOR = "IELTS Import"


def _style_header_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _comment(ws, cell_ref: str, text: str) -> None:
    c = Comment(text, _AUTHOR)
    c.width = 300
    c.height = 120
    ws[cell_ref].comment = c


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_test_info(ws) -> None:
    ws.sheet_properties.tabColor = _INFO_TAB

    headers = ["Field", "Value", "Notes"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    rows = [
        ("title", "", "Required. The test title shown to students."),
        ("description", "", "Optional. A brief description of the test."),
        ("type", "academic", "academic or general. Default: academic."),
    ]
    for r in rows:
        ws.append(r)

    _col_widths(ws, [18, 40, 50])
    ws.freeze_panes = "A2"

    _comment(ws, "B4", "Must be exactly 'academic' or 'general'.")


def _reading_headers() -> list[str]:
    return ["order", "group", "type", "question", "options", "answer", "instruction"]


def _example_reading_rows() -> list[tuple]:
    return [
        (
            1,
            1,
            "true_false_ng",
            "The club was founded in 1990.",
            "",
            "True",
            "Answer True, False, or Not Given.",
        ),
        (
            2,
            1,
            "true_false_ng",
            "The membership fee was raised recently.",
            "",
            "False",
            "",
        ),
        (
            3,
            2,
            "mcq",
            "What is the main purpose of the article?",
            "A. To inform;B. To persuade;C. To entertain;D. To describe",
            "B",
            "Choose the correct letter.",
        ),
        (
            4,
            3,
            "gap_fill",
            "The river flows through the city of ___.",
            "",
            "London",
            "Write NO MORE THAN ONE WORD.",
        ),
        (
            5,
            4,
            "matching",
            "Ancient Rome",
            "i. Architecture;ii. Philosophy;iii. Military;iv. Trade",
            "iii",
            "Match each civilisation to its greatest achievement.",
        ),
        (
            6,
            5,
            "matching_headings",
            "Paragraph A",
            "i. The origins of the tradition;ii. A shift in attitudes;iii. Modern-day challenges;iv. Early criticism;v. Future prospects",
            "ii",
            "The reading passage has several paragraphs. Choose the most suitable heading for each paragraph from the list.",
        ),
        (
            7,
            5,
            "matching_headings",
            "Paragraph B",
            "i. The origins of the tradition;ii. A shift in attitudes;iii. Modern-day challenges;iv. Early criticism;v. Future prospects",
            "iv",
            "",
        ),
        (
            8,
            6,
            "matching_information",
            "a reference to the financial impact of the discovery",
            "A;B;C;D;E;F;G",
            "C",
            "Match each statement with the correct section (A-G). NB: You may use any letter more than once.",
        ),
        (
            9,
            6,
            "matching_information",
            "an example of international cooperation",
            "A;B;C;D;E;F;G",
            "A",
            "",
        ),
        (
            10,
            7,
            "matching_features",
            "believes early exposure is essential",
            "A. Dr Patel;B. Prof. Chen;C. Dr Osei",
            "B",
            "Match each opinion with the correct researcher.",
        ),
        (
            11,
            7,
            "matching_features",
            "raises concerns about long-term effects",
            "A. Dr Patel;B. Prof. Chen;C. Dr Osei",
            "A",
            "",
        ),
    ]


def _sheet_reading(ws, num: int) -> None:
    ws.sheet_properties.tabColor = _READING_TAB

    ws["A1"] = (
        "Paste the full passage text here (A1). "
        "Questions start at row 3."
    )
    ws["A1"].font = Font(italic=True, color="555555")
    ws.row_dimensions[1].height = 30

    headers = _reading_headers()
    ws.append([""])  # row 2 blank separator
    ws.append(headers)
    _style_header_row(ws, 3, len(headers))

    for row in _example_reading_rows():
        ws.append(row)

    _col_widths(ws, [8, 8, 18, 60, 50, 15, 50])
    ws.freeze_panes = "A4"

    _comment(ws, "A3", "Order = position WITHIN the group (1, 2, 3…). Not the absolute IELTS Q number. Import auto-renumbers to 1..N per group.")
    _comment(ws, "B3", "Group number — questions with the same group number share an instruction block. Leave blank to auto-group by consecutive type.")
    _comment(ws, "E3", "Semicolon-separated options, e.g. A.xxx;B.yyy;C.zzz")
    _comment(ws, "F3", "For MCQ/matching: letter only (A, B, ii, etc.).\nFor gap_fill/sentence_completion/short_answer: the exact answer word(s); use semicolons for multiple variants (e.g. flightless;flightless parrot).\nFor true_false_ng: True / False / Not Given.\nFor yes_no_ng: Yes / No / Not Given.")
    _comment(ws, "C3", "Allowed types: mcq, true_false_ng, yes_no_ng, gap_fill, sentence_completion, short_answer, matching, matching_headings, matching_information, matching_features\n\nyes_no_ng: statement in 'question', answer = Yes / No / Not Given.\nsentence_completion: sentence with ____ in 'question', max_words in 'options' (default 3).\nshort_answer: question in 'question', max_words in 'options' (default 3).\nmatching_headings: paragraph labels in 'question', headings in 'options' (repeated per row).\nmatching_information: statements in 'question', section letters in 'options'.\nmatching_features: statements in 'question', people/places in 'options'.\nFor matching subtypes: answer = bare prefix only (e.g. 'iii', 'A').\n\nCompound types (table_completion / note_completion / form_completion / summary_completion) — create via UI editor, not Excel.")


def _sheet_writing(ws) -> None:
    ws.sheet_properties.tabColor = _WRITING_TAB

    headers = ["order", "task_number", "type", "prompt", "instruction", "min_words", "essay_type"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    rows = [
        (
            1,
            1,
            "task1",
            "The graph below shows the percentage of households with a computer in three countries between 2000 and 2020. Summarise the information by selecting and reporting the main features, and make comparisons where relevant.",
            "Write at least 150 words.",
            150,
            "",
        ),
        (
            2,
            2,
            "essay",
            "Some people think that universities should provide graduates with the knowledge and skills needed by employers. Others think that the true function of a university is to give access to knowledge for its own sake. Discuss both views and give your opinion.",
            "Write at least 250 words.",
            250,
            "discussion",
        ),
    ]
    for r in rows:
        ws.append(r)

    _col_widths(ws, [8, 14, 12, 80, 50, 12, 22])
    ws.freeze_panes = "A2"

    _comment(ws, "C1", "Allowed types: task1, essay")
    _comment(
        ws,
        "G1",
        "Task 2 only. Values: opinion, discussion, problem_solution, "
        "advantages_disadvantages, double_question. Aliases: op, disc, ps/prob, ad/advdis, dq/double. Leave blank for Task 1.",
    )


def _listening_headers() -> list[str]:
    return ["order", "group", "type", "question", "options", "answer", "part"]


def _example_listening_rows(part: int) -> list[tuple]:
    return [
        (
            1,
            1,
            "gap_fill",
            "The club's membership fee is ___ per year.",
            "",
            "£120",
            part,
        ),
        (
            2,
            1,
            "gap_fill",
            "The club was founded in ___.",
            "",
            "1987",
            part,
        ),
        (
            3,
            2,
            "mcq",
            "What does the speaker say about the new schedule?",
            "A. It starts earlier;B. It is the same;C. It ends later",
            "A",
            part,
        ),
    ]


def _sheet_listening(ws, part: int) -> None:
    ws.sheet_properties.tabColor = _LISTENING_TAB

    ws["A1"] = f"listening_part{part}.mp3"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = (
        f"[Optional audioscript for part {part} — "
        "shown to student after the test]"
    )
    ws["A2"].font = Font(italic=True, color="555555")

    ws.append([""])  # row 3 blank separator

    headers = _listening_headers()
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    for row in _example_listening_rows(part):
        ws.append(row)

    _col_widths(ws, [8, 8, 18, 60, 50, 15, 8])
    ws.freeze_panes = "A5"

    _comment(ws, "A1", "The filename of the MP3 file you will upload after import (e.g. part1.mp3).")
    _comment(ws, "A2", "Optional audioscript. Will be shown to student after the test.")
    _comment(ws, "A4", "Order = position WITHIN the group (1, 2, 3…). Not the absolute IELTS Q number. Import auto-renumbers to 1..N per group.")
    _comment(ws, "B4", "Group number — questions with the same group share an instruction block.")
    _comment(ws, "E4", "Semicolon-separated options, e.g. A.xxx;B.yyy;C.zzz")
    _comment(ws, "F4", "Letter (A, B, …) for MCQ, exact text for gap_fill.")
    _comment(ws, "C4", "Allowed types: mcq, gap_fill, sentence_completion, short_answer, matching, true_false_ng, matching_information, matching_features\nsentence_completion: sentence with ____ in 'question', max_words in 'options' (default 3).\nshort_answer: question in 'question', max_words in 'options' (default 3).\nFor matching_information/features: statement in 'question', options in 'options', answer = bare letter (A, B, ...).\n\nCompound types (table_completion / note_completion / form_completion / summary_completion) — create via UI editor, not Excel.")


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_template_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # Remove default sheet
    default = wb.active
    wb.remove(default)

    ws_info = wb.create_sheet("Test Info")
    _sheet_test_info(ws_info)

    for i in range(1, 4):
        ws = wb.create_sheet(f"Reading {i}")
        _sheet_reading(ws, i)

    ws_w = wb.create_sheet("Writing")
    _sheet_writing(ws_w)

    for i in range(1, 5):
        ws = wb.create_sheet(f"Listening {i}")
        _sheet_listening(ws, i)

    return wb


def workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
